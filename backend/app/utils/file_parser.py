"""
MÓDULO: PARSEO DE ARCHIVOS A TEXTO/MARKDOWN
=============================================

Utilidades para convertir diferentes tipos de archivos a texto plano o markdown.
Esto permite que el contenido de los archivos sea utilizado como contexto para
el chatbot con Gemini.

Soporta:
- PDF (.pdf)
- Word (.docx)
- Excel (.xlsx)
- PowerPoint (.pptx)
- Archivos de texto (.txt, .md, .csv)
- Código fuente (.py, .js, .java, etc.)
"""

import os
from flask import current_app
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook


def parse_pdf(filepath: str) -> str:
    """Parsea un archivo PDF a texto."""
    try:
        reader = PdfReader(filepath)
        text_parts = []
        for page_num, page in enumerate(reader.pages, 1):
            text = page.extract_text()
            if text.strip():
                text_parts.append(f"--- Página {page_num} ---\n{text}")
        return "\n\n".join(text_parts) if text_parts else ""
    except Exception as e:
        raise Exception(f"Error al parsear PDF: {str(e)}")


def parse_docx(filepath: str) -> str:
    """Parsea un archivo Word (.docx) a texto."""
    try:
        doc = Document(filepath)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n\n".join(paragraphs)
    except Exception as e:
        raise Exception(f"Error al parsear DOCX: {str(e)}")


def parse_xlsx(filepath: str) -> str:
    """Parsea un archivo Excel (.xlsx) a texto."""
    try:
        workbook = load_workbook(filepath, data_only=True)
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"## Hoja: {sheet_name}\n")

            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    text_parts.append(row_text)

            text_parts.append("")  # Línea en blanco entre hojas

        return "\n".join(text_parts)
    except Exception as e:
        raise Exception(f"Error al parsear XLSX: {str(e)}")


def parse_text_file(filepath: str) -> str:
    """Parsea archivos de texto plano."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return f.read()
    except UnicodeDecodeError:
        # Intentar con latin-1 si UTF-8 falla
        try:
            with open(filepath, 'r', encoding='latin-1') as f:
                return f.read()
        except Exception as e:
            raise Exception(f"Error al leer archivo de texto: {str(e)}")
    except Exception as e:
        raise Exception(f"Error al parsear archivo de texto: {str(e)}")


def parse_file_to_text(filepath: str) -> str:
    """
    Parsea un archivo a texto plano según su extensión.

    Args:
        filepath: Ruta absoluta al archivo a parsear

    Returns:
        str: Contenido del archivo en formato texto/markdown

    Raises:
        FileNotFoundError: Si el archivo no existe
        Exception: Si falla el parseo del archivo
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")

    ext = os.path.splitext(filepath)[1].lower()

    try:
        if ext == '.pdf':
            return parse_pdf(filepath)
        elif ext == '.docx':
            return parse_docx(filepath)
        elif ext == '.xlsx':
            return parse_xlsx(filepath)
        elif ext in ['.txt', '.md', '.markdown', '.csv', '.py', '.js', '.java',
                     '.cpp', '.c', '.html', '.css', '.json', '.xml', '.rst']:
            return parse_text_file(filepath)
        else:
            raise Exception(f"Tipo de archivo no soportado: {ext}")

    except Exception as e:
        current_app.logger.error(f"Error al parsear archivo {filepath}: {str(e)}")
        raise Exception(f"No se pudo parsear el archivo: {str(e)}")


def can_parse_file(filename: str) -> bool:
    """
    Verifica si un archivo puede ser parseado basándose en su extensión.

    Args:
        filename: Nombre del archivo con extensión

    Returns:
        bool: True si el archivo puede ser parseado
    """
    # Extensiones soportadas por nuestros parsers
    supported_extensions = {
        # Documentos
        '.pdf', '.docx',
        # Hojas de cálculo
        '.xlsx', '.csv',
        # Texto
        '.txt', '.md', '.markdown', '.rst',
        # Código
        '.py', '.js', '.java', '.cpp', '.c', '.html', '.css', '.json', '.xml',
    }

    ext = os.path.splitext(filename)[1].lower()
    return ext in supported_extensions


def get_file_type_category(filename: str) -> str:
    """
    Obtiene la categoría del tipo de archivo.

    Args:
        filename: Nombre del archivo

    Returns:
        str: Categoría ('document', 'spreadsheet', 'presentation', 'image', 'text', 'code', 'other')
    """
    ext = os.path.splitext(filename)[1].lower()

    document_exts = {'.pdf', '.doc', '.docx', '.odt'}
    spreadsheet_exts = {'.xls', '.xlsx', '.csv'}
    presentation_exts = {'.ppt', '.pptx'}
    image_exts = {'.png', '.jpg', '.jpeg', '.gif', '.bmp'}
    text_exts = {'.txt', '.md', '.markdown', '.rst'}
    code_exts = {'.py', '.js', '.java', '.cpp', '.c', '.html', '.css', '.json', '.xml'}

    if ext in document_exts:
        return 'document'
    elif ext in spreadsheet_exts:
        return 'spreadsheet'
    elif ext in presentation_exts:
        return 'presentation'
    elif ext in image_exts:
        return 'image'
    elif ext in text_exts:
        return 'text'
    elif ext in code_exts:
        return 'code'
    else:
        return 'other'


def estimate_token_count(text: str) -> int:
    """
    Estima el número de tokens en un texto.

    Usa una estimación aproximada: 1 token ≈ 4 caracteres (regla general para español).

    Args:
        text: Texto a analizar

    Returns:
        int: Número estimado de tokens
    """
    if not text:
        return 0
    return len(text) // 4


def truncate_text(text: str, max_tokens: int) -> str:
    """
    Trunca un texto a un máximo de tokens aproximados.

    Args:
        text: Texto a truncar
        max_tokens: Número máximo de tokens

    Returns:
        str: Texto truncado con indicador si fue cortado
    """
    if not text:
        return ""

    max_chars = max_tokens * 4  # Aproximación

    if len(text) <= max_chars:
        return text

    truncated = text[:max_chars]
    return truncated + "\n\n[... contenido truncado por límite de tokens ...]"
